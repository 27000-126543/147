import os
from typing import Dict, List, Any, Optional
from datetime import date, datetime
import pandas as pd
import numpy as np
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from utils.logger import logger
from config.settings import settings
from utils.helpers import round_float


class ExportEngine:
    def __init__(self):
        self.output_dir = os.path.join(settings.OUTPUT_DIR, "reports")
        os.makedirs(self.output_dir, exist_ok=True)
        
    def export_to_pdf(self, report_data: Dict[str, Any], output_filename: str) -> str:
        output_path = os.path.join(self.output_dir, output_filename)
        
        doc = SimpleDocTemplate(
            output_path,
            pagesize=landscape(A4),
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm
        )
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#2c3e50'),
            alignment=TA_CENTER,
            spaceAfter=20
        )
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#34495e'),
            alignment=TA_LEFT,
            spaceAfter=12,
            spaceBefore=12
        )
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#34495e'),
            spaceAfter=6
        )
        
        story = []
        
        title = report_data.get('title', '营销分析报告')
        report_date = report_data.get('report_date', datetime.now().strftime('%Y-%m-%d'))
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(f"报告日期: {report_date}", normal_style))
        story.append(Spacer(1, 0.5 * cm))
        
        summary = report_data.get('summary', {})
        if summary:
            story.append(Paragraph("一、执行摘要", subtitle_style))
            summary_data = [['指标', '数值', '同比变化']]
            for key, value in summary.items():
                if isinstance(value, dict):
                    summary_data.append([
                        value.get('label', key),
                        value.get('value', ''),
                        value.get('change', '')
                    ])
                else:
                    summary_data.append([key, str(value), ''])
            
            summary_table = Table(summary_data, colWidths=[6 * cm, 4 * cm, 4 * cm])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecf0f1')),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 0.5 * cm))
        
        charts = report_data.get('charts', [])
        if charts:
            story.append(Paragraph("二、图表分析", subtitle_style))
            for i, chart_path in enumerate(charts, 1):
                if os.path.exists(chart_path):
                    try:
                        img = Image(chart_path, width=18 * cm, height=12 * cm)
                        story.append(img)
                        story.append(Spacer(1, 0.3 * cm))
                        if i % 2 == 0:
                            story.append(PageBreak())
                    except Exception as e:
                        logger.warning(f"图表导入失败: {chart_path}, 错误: {e}")
        
        data_tables = report_data.get('data_tables', [])
        if data_tables:
            if not charts:
                story.append(Paragraph("二、详细数据", subtitle_style))
            else:
                story.append(Paragraph("三、详细数据", subtitle_style))
            
            for table_data in data_tables:
                table_title = table_data.get('title', '')
                table_rows = table_data.get('data', [])
                
                if table_title:
                    story.append(Paragraph(table_title, styles['Heading3']))
                
                if table_rows:
                    max_cols = max(len(row) for row in table_rows)
                    col_width = min(16 * cm / max_cols, 4 * cm)
                    pdf_table = Table(table_rows, colWidths=[col_width] * max_cols)
                    
                    table_style = TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ecc71')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ])
                    
                    for i in range(1, len(table_rows)):
                        if i % 2 == 0:
                            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa'))
                    
                    pdf_table.setStyle(table_style)
                    story.append(KeepTogether([pdf_table, Spacer(1, 0.3 * cm)]))
        
        recommendations = report_data.get('recommendations', [])
        if recommendations:
            story.append(Paragraph("四、优化建议", subtitle_style))
            for i, rec in enumerate(recommendations, 1):
                story.append(Paragraph(f"{i}. {rec}", normal_style))
        
        doc.build(story)
        logger.info(f"PDF报告已生成: {output_path}")
        return output_path
    
    def export_to_excel(self, report_data: Dict[str, Any], output_filename: str) -> str:
        output_path = os.path.join(self.output_dir, output_filename)
        
        wb = Workbook()
        wb.remove(wb.active)
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        data_tables = report_data.get('data_tables', [])
        
        for table_data in data_tables:
            sheet_name = table_data.get('sheet_name', table_data.get('title', '数据'))[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            rows = table_data.get('data', [])
            if not rows:
                continue
            
            for row_idx, row in enumerate(rows, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    
                    if row_idx == 1:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = center_alignment
                        if isinstance(value, (int, float)):
                            cell.number_format = '#,##0.00'
            
            for col_idx in range(1, len(rows[0]) + 1):
                max_length = 0
                for row in rows:
                    if len(row) >= col_idx:
                        value = str(row[col_idx - 1])
                        max_length = max(max_length, len(value))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 30)
            
            if len(rows) > 5 and sheet_name == '渠道效能':
                self._add_excel_chart(ws, rows)
        
        if 'summary' in report_data and report_data['summary']:
            ws_summary = wb.create_sheet(title="摘要", index=0)
            summary = report_data['summary']
            
            ws_summary.cell(row=1, column=1, value="指标")
            ws_summary.cell(row=1, column=2, value="数值")
            ws_summary.cell(row=1, column=3, value="同比变化")
            
            for col in range(1, 4):
                cell = ws_summary.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            row_idx = 2
            for key, value in summary.items():
                if isinstance(value, dict):
                    ws_summary.cell(row=row_idx, column=1, value=value.get('label', key))
                    ws_summary.cell(row=row_idx, column=2, value=value.get('value', ''))
                    ws_summary.cell(row=row_idx, column=3, value=value.get('change', ''))
                else:
                    ws_summary.cell(row=row_idx, column=1, value=key)
                    ws_summary.cell(row=row_idx, column=2, value=value)
                
                for col in range(1, 4):
                    cell = ws_summary.cell(row=row_idx, column=col)
                    cell.border = thin_border
                    cell.alignment = center_alignment
                row_idx += 1
            
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 15
        
        charts = report_data.get('charts', [])
        if charts:
            ws_charts = wb.create_sheet(title="图表")
            row_idx = 1
            for chart_path in charts:
                if os.path.exists(chart_path):
                    ws_charts.cell(row=row_idx, column=1, value=f"图表 {row_idx}")
                    ws_charts.cell(row=row_idx, column=1).font = header_font
                    row_idx += 2
        
        wb.save(output_path)
        logger.info(f"Excel报告已生成: {output_path}")
        return output_path
    
    def _add_excel_chart(self, ws, rows):
        data_rows = rows[1:]
        if len(data_rows) < 2:
            return
        
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "各渠道ROI对比"
        chart1.y_axis.title = "ROI"
        chart1.x_axis.title = "渠道"
        
        data = Reference(ws, min_col=5, min_row=1, max_row=len(rows), max_col=5)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows))
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        chart1.height = 10
        chart1.width = 20
        
        ws.add_chart(chart1, f"A{len(rows) + 3}")
        
        if len(rows[0]) >= 8:
            chart2 = LineChart()
            chart2.title = "各渠道ROAS对比"
            chart2.y_axis.title = "ROAS"
            chart2.x_axis.title = "渠道"
            
            data2 = Reference(ws, min_col=8, min_row=1, max_row=len(rows), max_col=8)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats)
            chart2.height = 10
            chart2.width = 20
            
            ws.add_chart(chart2, f"K{len(rows) + 3}")
    
    def export_raw_data(self, df: pd.DataFrame, output_filename: str) -> str:
        output_path = os.path.join(self.output_dir, output_filename)
        
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='原始数据')
            
            workbook = writer.book
            worksheet = writer.sheets['原始数据']
            
            header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col in range(len(df.columns)):
                cell = worksheet.cell(row=1, column=col + 1)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col_idx, col_name in enumerate(df.columns, 1):
                max_length = max(
                    df[col_name].astype(str).str.len().max(),
                    len(str(col_name))
                )
                worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        logger.info(f"原始数据已导出: {output_path}")
        return output_path
    
    def batch_export(self, report_data: Dict[str, Any], base_filename: str) -> Dict[str, str]:
        results = {}
        
        if 'pdf' in settings.EXPORT_FORMATS:
            pdf_path = self.export_to_pdf(
                report_data,
                f"{base_filename}.pdf"
            )
            results['pdf'] = pdf_path
        
        if 'xlsx' in settings.EXPORT_FORMATS:
            excel_path = self.export_to_excel(
                report_data,
                f"{base_filename}.xlsx"
            )
            results['xlsx'] = excel_path
        
        return results
